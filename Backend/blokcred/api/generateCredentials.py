import random
import os
from openpyxl import workbook
from openpyxl import load_workbook
from django.conf import settings


xl_filepath = os.path.join(settings.BASE_DIR, "api/Template.xlsx")



wb = load_workbook(xl_filepath)  
ws = wb['Sheet1']

rows = ws.iter_rows(min_row= 2, max_row = 3, min_col= 1, max_col= 3)


# def randompassword(stringLength=7):  # creates a random password
#     # lettersAndDigits = string.ascii_letters + string.digits
#     # return ''.join(random.choice(lettersAndDigits) for i in range(stringLength))



def username():  # creates a random username
    for a in rows:
        userId = a.value() + random.randint(10,99)

    return userId


def generatePassword():  # creates a random username
    
    for a in rows:
        userPassword = a.value
        userPassword = a.value(1,3) + b.value(1,2) + random.randint(10,99)

    return userPassword